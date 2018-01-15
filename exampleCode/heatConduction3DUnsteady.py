#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A course example of unsteady 3D heat conduction.

Constant propersties, contant temperature boundary, zero gradient (insulated) 
boundary, transient.

NOTE: USE OF SPARSE MATRIXES IN MatrixA

Created on Sat Jan 13 13:10:11 2018

@author: Antti Mikkonen, a.mikkonen@iki.fi
"""
import scipy as sp
from scipy import linalg
from matplotlib import pyplot as plt
import scipy.sparse as sparse
from scipy.sparse import linalg as splinalg
import openpyxl
from scipy import interpolate
from matplotlib import cm
from mpl_toolkits.mplot3d import Axes3D
import time

##############################################################################

#def write_excell(A,b,T=None):
#
#    wb = openpyxl.Workbook()
#    dest_filename = 'AbT.xlsx'
#
#    wsA = wb.active
#    wsA.title = "A"
#
#    for row in range(len(A)):
#        wsA.append(list(A[row]))
#
#    wsb = wb.create_sheet(title="b")
#    wsb.append(list(b))
#
##    wsT = wb.create_sheet(title="T")
##    for row in range(len(T)):
##        wsT.append(list(T[row]))
#
#    wb.save(filename = dest_filename)


class MatrixA(object):
    def __init__(self, n_tot, full):
        self.full = full
        if self.full:
            self.A = sp.zeros((n_tot,n_tot), dtype=sp.float_)
        else:
            self.l = []
            self.m = []
            self.val = []

    def add(self, l, m, val):
        if self.full:
            self.A[l,m] += val
        else:
            self.l.append(l)
            self.m.append(m)
            self.val.append(val)

    def finalize(self):
        if self.full:
            return self.A
        else:
            return sparse.csr_matrix(sparse.coo_matrix(
                                   (self.val, [self.l, self.m])
                                  ))
        
def solver(L, T_B, T, ks, rhoc, n, dt, t_max, constant_T_batch_coords, 
           full=False):
    
    nx = n[0]; ny = n[1]; nz = n[2]
    n_tot = n.prod()
    
    # Cell sizes
    dx = L / n
    
    # Coords of cell centers
    xyz = sp.zeros((3,nx,ny,nz), dtype=sp.float_)
    x = (sp.ones(nx)*dx[0]).cumsum()-dx[0]/2 
    y = (sp.ones(ny)*dx[1]).cumsum()-dx[1]/2
    z = (sp.ones(nz)*dx[2]).cumsum()-dx[2]/2
    for kz in range(nz):
        for ky in range(ny):
            for kx in range(nx):
                xyz[0][kx,ky,kz] = x[kx]
                xyz[1][kx,ky,kz] = y[ky]    
                xyz[2][kx,ky,kz] = z[kz]
    
    # Face area
    Ax = dx[1]*dx[2]
    Ay = dx[0]*dx[2]
    Az = dx[0]*dx[1]

    # Volume
    dV = dx.prod()
    
    # Current time (s)
    t = 0
    
    # Number of steps
    steps = int(sp.ceil(t_max/dt))
    
    # Coefficients
    aW  = ks*Ax/dx[0]
    aE  = ks*Ax/dx[0]

    aS  = ks*Ay/dx[1]
    aN  = ks*Ay/dx[1]

    aB  = ks*Az/dx[2]
    aT  = ks*Az/dx[2]

    aP0 = rhoc*dV/dt
    
    # Source vector
    bConstant = sp.zeros(n_tot, dtype=sp.float_)
    # Coefficient matrix
    A = MatrixA(n_tot, full)
    # Temperatures    
    Ts = sp.zeros((steps+1,n_tot), dtype=sp.float_)
    # Initial tempereture
    Ts[0,:] = T

    #################################################
    # Boundary indexes
    #################################################
    # x boundaries
    xMin = []
    for k in range(n_tot):
        if k % nx == 0:
            xMin.append(k)
    xMin = sp.array(xMin, dtype=sp.int_)
#    xMax = []
#    for k in range(n_tot):
#        if (k + 1) % nx == 0:
#            xMax.append(k)
#    xMax = sp.array(xMax, dtype=sp.int_)

    # y boundaries
    yMin = []
    for k in nx*ny*sp.arange(nz):
        for kx in range(nx):
            yMin.append(k+kx)
    yMin = sp.array(yMin, dtype=sp.int_)
    yMax = []
    for k in nx*ny*sp.arange(nz):
        for kx in range(nx):
            yMax.append(k+kx+nx*ny-nx)
    yMax = sp.array(yMax, dtype=sp.int_)

    # z boundaries
#    zMin = sp.arange(nx*ny, dtype=sp.int_)
#    zMax = sp.arange((nz-1)*nx*ny, (nz)*nx*ny, dtype=sp.int_)
    
    # yzMin
#    yzMin = sp.intersect1d(yMin,zMin)

    # Build inner diffusion
    for k in range(n_tot):

        # Add aW to matrix A
        if k % n[0] != 0:
            A.add(k,k, aW)
            A.add(k,k-1,-aW)

        # Add aE to matrix A
        if (k + 1) % n[0] != 0:
            A.add(k,k, aE)
            A.add(k,k+1,-aE)

        # Add aS to matrix A
        if k not in yMin:
            A.add(k,k, aS)
            A.add(k,k-nx,-aS)

        # Add aN to matrix A
        if k not in yMax:
            A.add(k,k, aN)
            A.add(k,k+nx,-aN)

        # Add aB to matrix A
        if k >= nx*ny:
            A.add(k,k, aB)
            A.add(k,k-nx*ny,-aB)
            
        # Add aT to matrix A
        if k < (nz-1)*nx*ny:
            A.add(k,k, aT)
            A.add(k,k+nx*ny,-aT)

    ########################################
    # Add time coefficient to diagonal
    ########################################
    for k in range(n_tot):
        A.add(k,k,aP0)
        
    ########################################
    # Constant T boundary
    ########################################
    cnY = constant_T_batch_coords[0] / dx[1]
    cnZ = constant_T_batch_coords[1] / dx[2]
    
    # z,y
    xMinMap = xMin.reshape((nz,ny))
    
    # Full faces
    cnYi = int(cnY)
    cnZi = int(cnZ)
    
    for ky in range(cnYi):
        for kz in range(cnZi):
            k = xMinMap[kz,ky]
            A.add(k,k,2*aW)
            bConstant[k] += 2*aW*T_B
            
    if cnYi < ny:
        for kz in range(cnZi):
            k = xMinMap[kz,cnYi]
            A.add(k,k,2*aW*cnY%1)
            bConstant[k] += 2*aW*T_B*cnY%1
    
    if cnZi < nz:
        for ky in range(cnYi):
            k = xMinMap[cnZi,ky]
            A.add(k,k,2*aW*cnZ%1)
            bConstant[k] += 2*aW*T_B*cnZ%1        


    if cnYi < ny and cnZi < nz:
        k = xMinMap[cnZi,cnYi]
        A.add(k,k,2*aW*(cnZ%1)*(cnY%1))
        bConstant[k] += 2*aW*T_B*(cnZ%1)*(cnY%1)
    
    ########################################
    # Solution
    ########################################
    A_final = A.finalize()
    for step in range(1,steps+1):
        b  = bConstant + aP0*T
        if full:
            T  = sp.linalg.solve(A_final,b)  
        else:
            T = splinalg.spsolve(A_final, b)
        Ts[step] = T
        t += dt


    ########################################
    # Post process
    ########################################
    T3d = sp.zeros((steps+1,nx,ny,nz), dtype=sp.float_)
    for step in range(1,steps+1):
        k = 0
        for kz in range(nz):
            for ky in range(ny):
                for kx in range(nx):
                    T3d[step,kx,ky,kz] = Ts[step][k]
                    k += 1
    
    return T3d, dx, xyz

    

##############################################################################

if __name__ == "__main__":
    def verify_with_1d():
        # Number of control volumes
        n = sp.array([10, 1, 1])
        # Lenght (m)
        L = sp.array([0.3, 0.5, 0.5])
        
        constant_T_batch_coords = L[1:]
        # Boundary temperatures (C)
        T_B   = 0
        # Initial temperatures (C)
        T_0 = 100
        # Thermal conductivity (W/mK)
        ks = 10 
        # Product of density and heat capasity (J/m3K)
        rhoc = 10e6
        # Time step (s)
        dt = 1
        # Stop time (s)
        t_max = 10
    
        Ts3d, dx3d, xyz = solver(L, T_B, T_0, ks, rhoc, n, dt, 
                                   t_max,constant_T_batch_coords,full=False)
    
        import heatConduction1DUnsteady
        Ts1d, dx1d = heatConduction1DUnsteady.solver(L[0], T_B, T_0, ks, rhoc, 
                                                     n[0], dt, t_max)
        
        T3d = Ts3d[-1,:,0,0]
        T1d = Ts1d[-1][::-1]
        print(T3d)
        print(T1d)
        assert(all(sp.isclose(T3d,T1d)))
        print("OK")


    ###########################################################################

    def main():
        
        
        # Lenght (m)
        L = sp.array([0.3, 0.4, 0.2])
        
        dx = 0.01
        dy = 0.05
        dz = 0.05
        
        # Number of control volumes
        n = sp.array([int(sp.ceil(L[0]/dx)), int(sp.ceil(L[1]/dy)), 
                      int(sp.ceil(L[2]/dz))])
        
        print("n", n)
        print("n tot", n.prod())
        
        constant_T_batch_coords = [1e-2,1e-2]
        
        # Boundary temperatures (C)
        T_B   = 100
        # Initial temperatures (C)
        T_0 = 0
        # Thermal conductivity (W/mK)
        ks = 10
        # Product of density and heat capasity (J/m3K)
        rhoc = 10e6
        # Time step (s)
        dt = 60*60
        # Stop time (s)
        t_max = 7*24*60*60
    
        # Plot at times (s)
        t_plot = sp.linspace(0,t_max, 15)#[0,40,80,t_max]
    
        Ts, dx, xyz = solver(L, T_B, T_0, ks, rhoc, n, dt, 
                            t_max,constant_T_batch_coords)
    
        plt.figure(0)
        x       = sp.linspace(dx[0]/2,L[0]-dx[0]/2,n[0])
        for t in t_plot:
            step = int(t/dt)
#            T = Ts[step,coords["yzMin"]]
            T = Ts[step,:,0,0]
            plt.plot(x*1e3, T, "d")
#            print(str(t).ljust(4), T)
        plt.xlim(0,L[0]*1e3)
        plt.xlabel("x (mm)")
        plt.ylabel("T (C)")
    
        # Plot first and last cell temperature
        plt.figure(1)
        t = sp.arange(0,t_max+1,dt)
        plt.plot(t, Ts[:,0,0,0], label="first")
        plt.plot(t, Ts[:,-1,0,0], label="last")
        plt.legend()
        plt.xlim(0,t_max)
        plt.xlabel("t (s)")
        plt.ylabel("T (C)")
    
    ###########################################################################

    def timeit():
        L = sp.array([0.3, 0.4, 0.2])
        constant_T_batch_coords = [1e-2,1e-2]
        
        # Boundary temperatures (C)
        T_B   = 100
        # Initial temperatures (C)
        T_0 = 0
        # Thermal conductivity (W/mK)
        ks = 10
        # Product of density and heat capasity (J/m3K)
        rhoc = 10e6
        # Time step (s)
        dt = 60*60
        # Stop time (s)
        t_max = 7*24*60*60
    
        full_t   = []
        sparse_t = []
    
        divs = sp.arange(1,10+1)
        ns   = sp.zeros(len(divs))
        for k in divs:
            print(k)
            dx = 0.3 / k
            dy = 0.4 / k
            dz = 0.2 / k
            
            # Number of control volumes
            n = sp.array([int(sp.ceil(L[0]/dx)), int(sp.ceil(L[1]/dy)), 
                          int(sp.ceil(L[2]/dz))])
        
            ns[k-1] = n.prod()
    
            start = time.time()
            Ts, dx, xyz = solver(L, T_B, T_0, ks, rhoc, n, dt, 
                                t_max,constant_T_batch_coords, full=False)
            sparse_t.append(time.time()-start)    
        
            start = time.time()
            Ts, dx, xyz = solver(L, T_B, T_0, ks, rhoc, n, dt, 
                                t_max,constant_T_batch_coords, full=True)
            full_t.append(time.time()-start)    
        
        full_t = sp.array(full_t)
        sparse_t = sp.array(sparse_t)
        
        print(full_t/sparse_t)
        
        plt.plot(ns,full_t/sparse_t,"d-k")
        plt.xlabel("cells")
        plt.ylabel("full time / sparse time")
        plt.savefig("fullVSsparseTime.pdf")
    
    ###########################################################################

    print("START")
#    verify_with_1d()
    main()
#    timeit()
    print("END")